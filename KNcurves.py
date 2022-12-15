from tkinter import filedialog as fd

from typing import Any, Dict, Generator

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils import rows_from_range, get_column_letter

import math
import os

from point2D import Point2D

'''***************** Excel Interface ****************'''

def loadSectionFromTable(ws:Worksheet, tb:Table)->dict:
    """Return a dcit (x) of lists, each list is an half cross section"""
    #{0: [Point2D(x,y), Point2D(x,y), Point2D(x,y), Point2D(x,y)], 30: [Point2D(x,y), Point2D(x,y), Point2D(x,y), Point2D(x,y)]}
    table = {}
    headers = tb.column_names
    #Check that the table contains colums x y z 
    if headers.count('x')*headers.count('y')*headers.count('z') == 0 :
        print("Table tbl_Hullform shall contain at leas columns 'x', 'y' and 'z'. Only the followings are detected:")
        print (tb.column_names)

    nX = tb.column_names.index('x')
    nY = tb.column_names.index('y')
    nZ = tb.column_names.index('z')

    #Removing header line
    dataBodyRange=ws[tb.ref]
    dataBodyRange = dataBodyRange[1:]
    #iter_rows = rows_from_range(tb.ref)

    for row_cells in dataBodyRange:
        if row_cells[nX].value not in table:
            table[row_cells[nX].value] = [Point2D(row_cells[nY].value,row_cells[nZ].value)]
        else:
            table[row_cells[nX].value].append(Point2D(row_cells[nY].value,row_cells[nZ].value))

    return table    

def getNamedRangeValue(wb:Workbook, rangeName:str):
    """Return the last value of a named range"""
    value = None
    dests = wb.defined_names[rangeName].destinations
    for sheet, coord in dests:
        value = wb[sheet][coord].value
    return value

def writeNotes(wb:Workbook,sheet:str,datas:list):
    #Delete the sheet if exists
    if sheet in wb.sheetnames:
        del wb[sheet]

    ws = wb.create_sheet(sheet) # insert at the end (default)
    
    for row in datas:
        ws.append(row)


def writeToWorkbook(wb:Workbook, table:list, name:str):
    '''Create a new sheet in the workbook, a table (tbl_name) and write the data which are list of dict. Header is the key of dict'''
    #Delete the sheet if exists
    if name in wb.sheetnames:
        del wb[name]

    ws = wb.create_sheet(name) # insert at the end (default)
    ws.sheet_properties.tabColor = 'C00000'

    # add column headings. NB. these must be strings
    ws.append(list(table[0].keys()))

    #Adding the datas
    for row in table:
        ws.append(list(row.values()))
    
    # The reference is needed, so we have to get it somehow.
    # Since the data is inserted starting in cell A1, all that remains 
    # is to get the last column and last row address:
    table_ref = 'A1:%s%d' % (get_column_letter(len(table[0])), len(table)+1)
    # Add the table to the worksheet:
    tab = Table(displayName='tbl_' + name, ref=table_ref)
   
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    
    ws.add_table(tab)

    #saveFile(filename,wb) #Need to save as a workaround preventing resetting the Autofilter
    #tab.autoFilter = None
    return tab  

def getAvailableFilename(f:str) -> str:
    '''Get the next filename available (not already open) using f + int'''
    baseName = os.path.splitext(os.path.basename(f))[0]
    extName = os.path.splitext(os.path.basename(f))[1]
    pathName = os.path.split(f)[0]

    fileIsAvailable = False
    n = 0
    while not fileIsAvailable:
       
        if n > 0:
            fname = pathName + '/' + baseName + str(n) + extName
        else:
            fname = f
        if os.path.exists(fname):
            try:
                #try to rename to check if it is open
                os.rename(fname, fname)
                fileIsAvailable = True
            except OSError as e:
                print ('File is already open by another app. Renaming', fname)
                n += 1
        else:
            fileIsAvailable = True
    return fname


def loadFromExcel(filename)->dict:
    print ('Loading datas from',filename)

    wb = load_workbook(filename, data_only=True)

    #Load data from table tbl_Hullform
    wsInput = wb['Input']
    tblHullForm = wsInput.tables["tbl_Hullform"]

    hullDatas = loadSectionFromTable(wsInput, tblHullForm)

    #Read constant values in excel file
    maxWl = getNamedRangeValue(wb,'max_wl')
    deltaWl = getNamedRangeValue(wb,'Δwl')
    maxAngle = getNamedRangeValue(wb,'φMax')
    deltaAngle = getNamedRangeValue(wb,'Δφ')
    d_sw = getNamedRangeValue(wb,'ρsw')

    wb.close()

    return {'hullDatas':hullDatas, 'maxWl':maxWl, 'deltaWl':deltaWl, 'maxAngle':maxAngle, 'deltaAngle':deltaAngle, 'd_sw':d_sw}

def saveToExcel(f:str, tables:dict):
    
    fname = getAvailableFilename(f)
    wb = load_workbook(f)

    writeNotes(wb, 'Notes', notes({'d_sw':tables['d_sw']}))

    tab1 = writeToWorkbook(wb, tables['hydrotable'],'HydroComputation')
    tab2 = writeToWorkbook(wb, tables['KNdatas'],'KNComputation')
    tab3 = writeToWorkbook(wb, tables['KNobject'],'KNTable')

    #Workaround for the bug that autofilter are erased on tables when saving the 1st time
    wb.save(fname)
    tab1.autoFilter = None
    tab2.autoFilter = None
    tab3.autoFilter = None
    wb.save(fname)
    print ('----------- Saved in',fname,'----------')
    wb.close()


''' ****************** Core computation **************** '''

def computeInter(segment:list, hydroCondition:dict)->Point2D:
    """Compute Intersection of segment and wl. Return None if no intersection"""
    ptA = segment[0]
    ptB = segment[1]
    ptC = Point2D(0,hydroCondition['waterline'])
    xD = max(ptA.x,ptB.x)+1 #TODO : xD = 0 ?????
    ptD = Point2D(xD,hydroCondition['waterline'] + xD * math.tan(hydroCondition['Phi'] * math.pi / 180) )

    vAB = ptB - ptA
    vWL = ptD - ptC
    div = vAB.x * vWL.y - vAB.y * vWL.x
    m = 0.0
    k = 0.0
    if (div != 0):
        m = (vAB.x * ptA.y - vAB.x * ptC.y - vAB.y * ptA.x + vAB.y * ptC.x ) / div
        k = (vWL.x * ptA.y - vWL.x * ptC.y - vWL.y * ptA.x + vWL.y * ptC.x ) / div
        if (0.0<=m) and (m<1) and (0<=k) and (k<1):
            #there is an intersection
            ptInt = ptA +  vAB * k
            return ptInt
    #No intersection
    return None

def computeZandS(segment:list, hydroCondition:dict)->dict:
    """Compute CoG and surface of the trapeze from hyrdo to segment"""
    smallBase = hydroCondition['waterline'] - max(segment[0].y, segment[1].y)
    bigBase = hydroCondition['waterline'] - min(segment[0].y, segment[1].y)
    height = segment[1].x - segment[0].x
    areaRect = smallBase * height
    zCoGRect = hydroCondition['waterline'] - smallBase / 2
    areaTriangle = (bigBase - smallBase) * height / 2 
    zCoGTriangle = min(segment[0].y,segment[1].y) + (bigBase-smallBase)* 2/3

    Area = areaRect + areaTriangle

    if not math.isclose(Area,0):
        Zcog = (areaRect * zCoGRect +  areaTriangle * zCoGTriangle) / Area
    else:
        Zcog = 0

    return {'Area':Area,'Zcog':Zcog}
    


def segmentList(section:list)->list:
    """From a list of point, create a list of pair (list) of segements."""
    segList =[]
    for a,b in zip(section[:-1],section[1:]):
        segList.append([a,b])
    return segList



def computeZc(section:list, hydroCondition:dict)->dict:
    """Compute Zc and area below wl for a given Hull section."""
    trapezeList=[]
    waterHalfWidth = 0.0
    for segment in segmentList(section):
        if (min(segment[0].y,segment[1].y)<hydroCondition['waterline'])and not math.isclose(min(segment[0].y,segment[1].y),hydroCondition['waterline']):
            ptInt = computeInter(segment, hydroCondition)
            if not (ptInt is None):
                if math.isclose(waterHalfWidth,0):
                    waterHalfWidth = ptInt.x
                else:
                    waterHalfWidth = ptInt.x - waterHalfWidth
                if (segment[0].y<ptInt.y) or math.isclose(segment[0].y,ptInt.y):
                    trapezeList.append([segment[0], ptInt])
                else:
                    trapezeList.append([ptInt, segment[1]])
            else:
                trapezeList.append(segment)

    Area = 0
    Mz = 0
    for segment in trapezeList:
        datas = computeZandS(segment, hydroCondition)
        Area += datas['Area']
        Mz += datas['Area'] * datas['Zcog']

    Zc = Mz / Area
    return {'Zc':Zc,'Area':Area, 'WaterHalfWidth':waterHalfWidth}

    
    
def computeCenterOfBuyoancy(hull:dict,wl:float,d_sw:float)->dict:
    """Compute VCB and Volume below wl for a given Waterline."""
    hydroCondition = {'waterline':wl,'Phi':0.0}
    Mx = 0.0
    Mz = 0.0
    Volume = 0.0
    WaterplaneArea = 0.0
    RMT = 0.0
    RML = 0.0
    Lpp = 0.0
    nLCF =0
    LCF = 0.0
    for index, x in enumerate(hull):
        #the length of the last element will be the same as the n-1 one
        if index < (len(hull) - 1):
            elmntLength = list(hull)[index+1] - list(hull)[index]
        
        datas = computeZc(hull[x],hydroCondition)

        eltVolume = datas['Area'] * elmntLength
        xelt = list(hull)[index] + elmntLength / 2
        Mx += xelt *  eltVolume
        Mz += datas['Zc'] * eltVolume
        Volume += eltVolume
        WaterplaneArea += datas['WaterHalfWidth'] * 2 * elmntLength
        RMT += elmntLength * (2 * datas['WaterHalfWidth'])**3 /12 # TODO:This computation don't work for a catamaran (should transport this result)
        RML += (2 * datas['WaterHalfWidth']) * elmntLength**3  /12 + ((2 * datas['WaterHalfWidth']) * elmntLength) * xelt**2 # RML at x=0
        Lpp += elmntLength

        if not math.isclose(datas['WaterHalfWidth'],0):
            #part of this section is underwater
            if nLCF==0: 
                LCF += x
                nLCF += 1
            LCF += x + elmntLength
            nLCF += 1
            
    
    LCF /= nLCF
    LCB = Mx / Volume
    VCB = Mz / Volume
    Volume *=2 #half section given
    Displacement = Volume * d_sw
    Immersion =  WaterplaneArea * d_sw / 100 #in t/cm
    RMT /= Displacement
    RML -= WaterplaneArea * LCB**2 #transport RML to CoB
    RML /= Displacement
    MCT = Displacement * RML / (100*Lpp)
    KMT = RMT + VCB
    return {'waterline' : wl,'Volume': Volume,'Displacement': Displacement,'Immersion': Immersion,'MCT': MCT,'LCB':LCB,'TCB':0.0, 'LCF': LCF,'KMT': KMT,
            'WaterplaneArea': WaterplaneArea,'RMT': RMT,'RML': RML,'Lpp': Lpp,'VCB':VCB}


def computeHydroTable(hull:dict,deltaWl:float,maxWl:float,d_sw:float):
    """Compute Volume, LCB, VCB for each waterline step.Return a list of dict"""
    Hydrotable =[]
    wl = deltaWl
    while wl <= maxWl:
        Hydrotable.append(computeCenterOfBuyoancy(hull,wl,d_sw))
        wl += deltaWl
        

    return Hydrotable

def computeDist(section:list, hydroCondition:dict)->float:
    """Compute y (waterline length) at a given angle and wl for a given section."""
    distMax = 0.0
    A = Point2D(0,hydroCondition['waterline'])
    
    for segment in segmentList(section):
        ptInter = computeInter(segment, hydroCondition)
        if not (ptInter is None):
            distMax = max(A.distanceTo(ptInter),distMax)
    
    return distMax


def computeKNatAngle(hull:dict,hydrotable:list,angle:float)->list:
    """Compute C0C1, d, Hmeta, KNsin for each waterline step at a given angle.Return a list of dict"""

    KNdatas=[]
    for wl_data in hydrotable:
        
        hydroPhiPlus = {'waterline':wl_data['waterline'],'Phi': angle / 2}
        hydroPhiMinus = {'waterline':wl_data['waterline'],'Phi': - angle / 2}

        yp = []
        ym = []
        yp2 = []
        ym2 = []
        yp3 = []
        ym3 = []
        for index, x in enumerate(hull):
            #the length of the last element will be the same as the n-1 one
            if index < (len(hull) - 1):
                elmntLength = list(hull)[index+1] - list(hull)[index]
            _yp = computeDist(hull[x],hydroPhiPlus)
            _ym = computeDist(hull[x],hydroPhiMinus)
            yp.append(_yp)
            ym.append(_ym)
            yp2.append(_yp**2)
            ym2.append(_ym**2)
            yp3.append(_yp**3)
            ym3.append(_ym**3)

        VCB = wl_data['VCB']
        LCB = wl_data['LCB']
        Im = elmntLength/3*(sum(yp3)+sum(ym3))-elmntLength/4*((sum(yp2)-sum(ym2))**2)/(sum(yp)+sum(ym))
        C0C1 = 2 * Im / wl_data['Volume'] * math.sin ( angle / 2 * math.pi /180 )
        C0Cperp = C0C1 * math.cos( angle / 2 * math.pi /180 )
        d = 0.5 * (sum(yp2)-sum(ym2)) / (sum(yp)+sum(ym))
        Hmeta = C0Cperp / math.sin ( angle * math.pi /180 )
        KNsin = (Hmeta + VCB) * math.sin ( angle * math.pi /180 )

        KNrow = {'waterline':wl_data['waterline'],'Volume':wl_data['Volume'],'Displacement': wl_data['Displacement'],
                    'Phi':angle,'C0C1':C0C1, 'd':d,'Hmeta':Hmeta,'KNsin':KNsin}
        KNdatas.append(KNrow)
        
    return KNdatas


def computeKNdatas(hull:dict,hydrotable:list,deltaAngle:float,maxAngle:float)->list:
    """Compute Volume, Zc, Hmeta, KN for each waterline step and each angle.Return a list of dict"""
    KNdatas =[]
    angle = 0.00001
    while angle <= maxAngle:
        KNdatas.extend(computeKNatAngle(hull,hydrotable,angle))
        angle += deltaAngle 

    return KNdatas
    
def formatKNtables(KNdatas:dict)->dict:
    """Compute KN table: KNsin for each waterline and a column for each phi"""
    #Extract list of angles without duplicates
    KNtable = []
    PhiList = list(dict.fromkeys([d['Phi'] for d in KNdatas if 'Phi' in d]))
    wlList = list(dict.fromkeys([d['waterline'] for d in KNdatas if 'waterline' in d]))
    
    RowHeader = [['waterline','Volume','Displacement']]

    for wl in wlList:
        wlRow = {}
        wlDatas = [r for r in KNdatas if math.isclose(r['waterline'], wl)]

        wlRow['waterline'] = wlDatas[0]['waterline']
        wlRow['Volume'] = wlDatas[0]['Volume']
        wlRow['Displacement'] = wlDatas[0]['Displacement']

        for Phi in PhiList:
            KNvalue = [d['KNsin'] for d in wlDatas if math.isclose(d['Phi'], Phi)]
            wlRow['%.2f' % Phi] = KNvalue[0]
        
        KNtable.append(wlRow)

    return KNtable


def computeDatas(datas:dict)->dict:
    
    print('Starting Hydrostatic Computation with', len(datas['hullDatas']), 'cross sections from', datas['deltaWl'],'m to',datas['maxWl'],'m')
    hydrotable = computeHydroTable(datas['hullDatas'],datas['deltaWl'],datas['maxWl'],datas['d_sw'])

    print('*** Starting KN Computation from 0° to', datas['maxAngle'],'° steps of',datas['deltaAngle'],'° ***')
    KNdatas = computeKNdatas(datas['hullDatas'],hydrotable,datas['deltaAngle'],datas['maxAngle'])

    KNobject = formatKNtables(KNdatas)

    return {'hydrotable':hydrotable, 'KNdatas':KNdatas, 'KNobject':KNobject, 'd_sw':datas['d_sw'] }

def notes(args:dict)->list:
    datas = [
    ['Hydrostatics'],
    ['Data', 'Unit', 'Comment'],
    ['waterline', 'm','water height from keel + upward'],
    ['Volume', 'm3','Immerged hull volume'],
    ['Displacement', 't','Displacement considering density of ' + str(args['d_sw'])],
    ['Immersion', 't/cm','Weight to sink 1cm'],
    ['MCT', 't.m/cm','Moment to Change Trim by 1cm'],
    ['LCB', 'm','Longitudinal position of Center Of Buoyancy from Aft'],
    ['TCB', 'm','Transversal position of Center Of Buoyancy from centerline, always 0'],
    ['LCF', 'm','Longitudinal position of Center Of Flotation from Aft'],
    ['KMT', 'm','Transversal Metacentric Height above Keel'],
    ['WaterplaneArea', 'm2','Area of the hull @ waterline'],
    ['RML', 'm','Longitudinal Metacentric Radius above keel'],
    ['RMT', 'm','Transversal Metacentric Radius above keel'],
    ['Lpp', 'm','Length between perpendicular'],
    ['VCB', 'm','Vertical position of Center Of Buoyancy from Keel'],
    ['-----'],
    ['KNcomputation'],
    ['Data', 'Unit', 'Comment'],
    ['waterline', 'm','water height from keel + upward'],
    ['Volume', 'm3','Immerged hull volume'],
    ['Displacement', 't','Displacement considering density of ' + str(args['d_sw'])],
    ['Phi', '°','List angle'],
    ['C0C1', 'm','Distance between CoB at null list and CoB at list angle (at phi/2)'],
    ['d', 'm','x position of the intersecion waterline at lista angle with horizontal waterline'],
    ['Hmeta', 'm','Transversal Metacentric Height above CoB'],
    ['KNsin', 'm','Projected distance of K on the inclined centerline'],
    ]

    return datas

# *********** Entry point ***************
print('Workbook shall contain a sheet "Input" with the named range:')
print('tbl_Hullform (half hull), max_wl, Δwl, φMax, Δφ, ρsw')
print('All data shall be x: longitudinal forward y: transveral starboard z: vertical upward')
print('x=0: aft perpendicular y=0: centerline z=0: keel')

filename = fd.askopenfilename(filetypes=[("Excel files", ".xlsx .xls")])
if not filename:
    print('No file was selected. Exiting')
    quit()

print('---------- Starting ----------')
datas = loadFromExcel(filename)
tables = computeDatas(datas)
saveToExcel(filename, tables)














